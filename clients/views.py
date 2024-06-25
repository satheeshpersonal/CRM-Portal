from django.shortcuts import render
from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from rest_framework.authentication import TokenAuthentication, SessionAuthentication, BasicAuthentication
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.contrib.auth import authenticate
from django.db.models import Q, Count
from django.template.defaultfilters import slugify

from openpyxl import load_workbook

from .serializer import *
from .models import *
from siteconfig.functions import get_validation_error_message, method_not_allowed

import environ
env=environ.Env()

# Create your views here.

class DocumentView(viewsets.ModelViewSet):
    serializer_class = ClientSerializer
    authentication_classes = [TokenAuthentication, SessionAuthentication, BasicAuthentication]
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs): # Create new client by documnet 
        try:
        # if 1==1:
            
            if hasattr(request.data,'_mutable'):
                request.data._mutable=True
            request.data["created_by"] = request.data["updated_by"] = request.user.id

            doc_serializer =DocumentSerializer(data=request.data)
            if doc_serializer.is_valid():
                document = doc_serializer.save()
            else:
                error_message = get_validation_error_message(doc_serializer)
                return Response({"status" : "failure", "data":{}, "error": {"error_message":error_message ,"error_code":"error"}, "extra_data" : {}}, status=400)
            

            excel_file = request.FILES['document_file']
            wb = load_workbook(excel_file)
            ws = wb.active

            table_head = []
            index = 0
            for row in ws.iter_rows(min_row=0, values_only=True):
                if index == 0:
                    for heare_val in row:
                        table_head.append(slugify(heare_val).replace("-", "_"))
                    print("table_head -- ", table_head)
                else:
                    # name, price, quantity = row
                    if row[0]:
                        # print("row ----- ", row)
                        insert_row = {}
                        row_index = 0
                        for attr in table_head:
                            insert_row[attr]=row[row_index]
                            row_index +=1

                        insert_row["created_by"] = insert_row["updated_by"] = request.user.id
                        insert_row["document"] = document.id

                        try:
                            # clients, created = Clients.objects.update_or_create(
                            #     adresse_e_mail=insert_row["adresse_e_mail"],
                            #     defaults={insert_row})

                            print(insert_row)
                            serializer = self.get_serializer(data=insert_row)
                            if serializer.is_valid():
                                client_data = serializer.save()

                                client_access = {"user":request.user.id, "clients":client_data.id}
                                
                                client_access_serializer = ClientAccessSerializer(data=client_access)
                                if client_access_serializer.is_valid():
                                    client_access_serializer.save()
                                else:
                                    error_message = get_validation_error_message(client_access_serializer)
                                    print("add client access-- ", error_message)

                            else:
                                error_message = get_validation_error_message(serializer)
                                print("add client -- ", error_message)
                
                            # horodateur,adresse_email,nom_postnom_et_prenom,lieu_date_de_naissance,telephone,profession,etat_civil,conjoint,date_benediction_nuptiale,nombre_denfants,noms_prenoms_enfants,date_de_consecretion,niveau_detude,adresse_locale,province,ville,pays,nationalite,communion,coordonnateur_communion,coordonnateur_asst_communion,sec_tresorier_co,munion,adresse_communion,telephone_coordonnateur_communion,rameau,coordonnateur_rameau,coordonnateur_assist_ram,au,sec_tresorier_rameau,secteur,coordonnateur_secteur,secretaire_secteur,section,coordonnateur_de_section,secre,aire_section,date_1er_cours_fondamental_fr,date_1er_cours_intermediaire_fr,date_1er_cours_vivre_la_parole_de_di,u_en_famille,date_cours_fondamentale_en_anglais,date_cours_defeating_the_adversary,date_cours_living_the_book_o,_acts,date_cours_avance,date_cours_renewing_the_mind,date_cours_practical_keys,date_cours_living,fonction_au_mi,istere = row
                        except Exception as e:
                            print("create client error - ",e) 
                    else:
                        print("value missing")

                index +=1

            return Response({"status" : "success", "data": {}, "error": {"error_message":"Document added successfully.","error_code":"success"}, "extra_data" : {}},status=200)
       
        except ValidationError as ve:
            return Response({"status" : "failure", "data": {}, "error": {"error_message":ve.messages[0],"error_code":"error"}, "extra_data" : {}}, status=400)            
        except Exception as e:
            print("exception - ",e) 
            return Response({"status" : "failure", "data": {}, "error": {"error_message":"Something went wrong", "error_code":"error"}, "extra_data" : {}}, status=400)  
    

    def list(self, request, *args, **kwargs): # get document list
        extra_data={}
        try:
            all_doc = Document.objects.filter(created_by = request.user.id).order_by("-created")
            # print("all_doc --------", all_doc)
            serializer_data = DocumentSerializer(all_doc, many=True).data 

            extra_data["base_url"] = env.str('BASE_URL')
            
            return Response({"status" : "success", "data": serializer_data, "error": {"error_message":"success.","error_code":"success"}, "extra_data" : extra_data},status=200)

        except ValidationError as ve:
            return Response({"status" : "failure", "data": {}, "error": {"error_message":ve.messages[0],"error_code":"error"}, "extra_data" : {}}, status=400)            
        except Exception as e:
            print("exception - ",e) 
            return Response({"status" : "failure", "data": {}, "error": {"error_message":"Something went wrong", "error_code":"error"}, "extra_data" : {}}, status=400)  
    

class ClientsAPIView(APIView):
    authentication_classes = [TokenAuthentication, SessionAuthentication, BasicAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        output_data = []
        # request.data

        all_clients = ClientAccess.objects.filter(user_id = request.user.id)
        # print(get_clients.query)
        print("get_clients  --------------- ", request.user.id)
        
        for get_clients in all_clients:
            # print("get_clients  --------------- ", get_clients)
            output_data.append(ClientSerializer(get_clients.clients).data)

        extra_data = {
        }
        
        return Response({"status" : "success", "data": output_data, "error": {"error_message":"success","error_code":"success"}, "extra_data" : extra_data},status=200)
